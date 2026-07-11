import openpyxl

def get_data():
    workbook = openpyxl.load_workbook('C://Selenium_E26/Flight_Framework_Mock/testdata/testdata.xlsx')
    sheet = workbook["login"]

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    return data

def get_data2():
    workbook = openpyxl.load_workbook('C://Selenium_E26/Flight_Framework_Mock/testdata/testdata.xlsx')
    sheet = workbook["stations"]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data